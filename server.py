"""
DataMapper — Agentic Source-to-Target Mapping Engine
FastAPI backend: schema upload → BQ crawl → LLM mapping → editable table → export

Loads .env from ../sql gen/.env (DeepSeek + BQ creds) then layers user-supplied
keys on top. Also supports Claude (anthropic SDK) as the default provider.
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
import os
import re
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, AsyncIterator, Dict, List, Optional

import anthropic
import httpx
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import (
    FileResponse,
    JSONResponse,
    StreamingResponse,
)
from fastapi.staticfiles import StaticFiles
from openai import OpenAI
from pydantic import BaseModel

# ── Load env from sql gen sibling folder ─────────────────────────────────────
_ENV_PATH = Path(__file__).parent.parent / "sql gen" / ".env"
if _ENV_PATH.exists():
    load_dotenv(_ENV_PATH, override=False)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s %(message)s")
logger = logging.getLogger("datamapper")

# ── Global config (overridable per-session via user API keys) ──────────────
_DEFAULT_PROVIDER   = os.getenv("DM_PROVIDER", "claude")        # claude | deepseek | custom
_ANTHROPIC_API_KEY  = os.getenv("ANTHROPIC_API_KEY", "")
_DEEPSEEK_API_KEY   = os.getenv("LLM_API_KEY", "")
_DEEPSEEK_BASE_URL  = os.getenv("LLM_BASE_URL", "https://api.deepseek.com/v1")
_DEEPSEEK_MODEL     = os.getenv("LLM_MODEL", "deepseek-chat")
_CLAUDE_MODEL       = os.getenv("DM_CLAUDE_MODEL", "claude-sonnet-4-6")
_BQ_PROJECT         = os.getenv("BQ_PROJECT_ID", "")
_BQ_DATASET         = os.getenv("BQ_DATASET", "")
_GCP_CREDS          = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "")
_JIRA_URL           = os.getenv("JIRA_URL", "")
_JIRA_EMAIL         = os.getenv("JIRA_EMAIL", "")
_JIRA_TOKEN         = os.getenv("JIRA_TOKEN", "")

# ── In-memory session store ────────────────────────────────────────────────
_sessions: Dict[str, Dict[str, Any]] = {}
_sse_queues: Dict[str, asyncio.Queue] = {}


# ─────────────────────────────────────────────────────────────────────────────
# LLM CLIENT — supports Claude (native) + any OpenAI-compat provider
# ─────────────────────────────────────────────────────────────────────────────

class MultiLLMClient:
    """Unified client: provider=claude uses anthropic SDK; others use openai SDK."""

    def __init__(
        self,
        provider: str = "claude",
        api_key: str = "",
        base_url: str = "",
        model: str = "",
    ):
        self.provider = provider.lower()
        self.model    = model or (_CLAUDE_MODEL if provider == "claude" else _DEEPSEEK_MODEL)

        if self.provider == "claude":
            key = api_key or _ANTHROPIC_API_KEY
            if not key:
                raise ValueError("Anthropic API key not configured. Add ANTHROPIC_API_KEY to .env or enter it in Settings.")
            self._anthropic = anthropic.Anthropic(api_key=key)
            self._openai    = None
        else:
            key  = api_key or _DEEPSEEK_API_KEY
            url  = base_url or _DEEPSEEK_BASE_URL
            self._openai    = OpenAI(api_key=key, base_url=url)
            self._anthropic = None

    def complete(self, system: str, prompt: str, temperature: float = 0.1, max_tokens: int = 4096) -> str:
        if self.provider == "claude":
            resp = self._anthropic.messages.create(
                model=self.model,
                max_tokens=max_tokens,
                temperature=temperature,
                system=system,
                messages=[{"role": "user", "content": prompt}],
            )
            return resp.content[0].text.strip()
        else:
            msgs = []
            if system:
                msgs.append({"role": "system", "content": system})
            msgs.append({"role": "user", "content": prompt})
            resp = self._openai.chat.completions.create(
                model=self.model,
                messages=msgs,
                temperature=temperature,
                max_tokens=max_tokens,
            )
            return (resp.choices[0].message.content or "").strip()

    def complete_json(self, system: str, prompt: str) -> Any:
        raw = self.complete(system, prompt)
        # Strip markdown fences
        m = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw)
        if m:
            raw = m.group(1).strip()
        # Try direct parse
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            m2 = re.search(r"(\{[\s\S]*\}|\[[\s\S]*\])", raw)
            if m2:
                return json.loads(m2.group(1))
            raise ValueError(f"LLM did not return valid JSON:\n{raw[:400]}")


def _make_llm(session: Dict) -> MultiLLMClient:
    """Build LLM client from session-level key overrides or global env."""
    cfg = session.get("api_config", {})
    provider = cfg.get("provider") or _DEFAULT_PROVIDER
    return MultiLLMClient(
        provider=provider,
        api_key=cfg.get("api_key", ""),
        base_url=cfg.get("base_url", ""),
        model=cfg.get("model", ""),
    )


# ─────────────────────────────────────────────────────────────────────────────
# SCHEMA PARSER — Excel / CSV → List[{name, type, sample, nullable}]
# ─────────────────────────────────────────────────────────────────────────────

def parse_schema_file(content: bytes, filename: str) -> Dict[str, Any]:
    """Return {tables: [{name, columns:[{name,type,sample,nullable}]}]}"""
    ext = filename.rsplit(".", 1)[-1].lower()

    if ext in ("xlsx", "xls"):
        return _parse_excel(content)
    elif ext == "csv":
        return _parse_csv(content)
    else:
        raise ValueError(f"Unsupported file type: .{ext}")


def _normalize_type(raw: str) -> str:
    raw = str(raw).upper().strip()
    MAP = {
        "INT": "INT64", "INTEGER": "INT64", "BIGINT": "INT64", "SMALLINT": "INT64",
        "FLOAT": "FLOAT64", "DOUBLE": "FLOAT64", "REAL": "FLOAT64", "DECIMAL": "NUMERIC",
        "BOOL": "BOOLEAN", "BIT": "BOOLEAN",
        "VARCHAR": "STRING", "NVARCHAR": "STRING", "CHAR": "STRING", "TEXT": "STRING",
        "DATE": "DATE", "DATETIME": "TIMESTAMP", "TIMESTAMP": "TIMESTAMP",
        "BLOB": "BYTES", "BINARY": "BYTES",
    }
    return MAP.get(raw.split("(")[0], raw or "STRING")


def _parse_excel(content: bytes) -> Dict[str, Any]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    tables = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c else "" for c in rows[0]]
        col_idx = {h.lower(): i for i, h in enumerate(header)}

        # Detect column-name column: first non-empty header
        name_col = next((col_idx.get(k) for k in ("field", "column", "column_name", "field_name", "name") if col_idx.get(k) is not None), 0)
        type_col = next((col_idx.get(k) for k in ("type", "data_type", "datatype", "dtype") if col_idx.get(k) is not None), None)
        sample_col = next((col_idx.get(k) for k in ("sample", "example", "sample_value") if col_idx.get(k) is not None), None)
        nullable_col = next((col_idx.get(k) for k in ("nullable", "null", "required") if col_idx.get(k) is not None), None)

        columns = []
        for row in rows[1:]:
            name = str(row[name_col]).strip() if row[name_col] else ""
            if not name or name.lower() in ("none", "nan", ""):
                continue
            col = {
                "name": name,
                "type": _normalize_type(str(row[type_col]).strip() if type_col is not None and row[type_col] else "STRING"),
                "sample": str(row[sample_col])[:100] if sample_col is not None and row[sample_col] else "",
                "nullable": str(row[nullable_col]).upper() not in ("NO", "NOT NULL", "FALSE", "0") if nullable_col is not None else True,
            }
            columns.append(col)
        if columns:
            tables.append({"name": sheet_name, "columns": columns})
    return {"tables": tables}


def _parse_csv(content: bytes) -> Dict[str, Any]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    header = [h.lower().strip() for h in (reader.fieldnames or [])]

    name_col = next((h for h in header if h in ("field", "column", "column_name", "field_name", "name")), header[0] if header else "name")
    type_col = next((h for h in header if h in ("type", "data_type", "datatype")), None)
    sample_col = next((h for h in header if h in ("sample", "example", "sample_value")), None)
    nullable_col = next((h for h in header if h in ("nullable", "null", "required")), None)

    columns = []
    for row in reader:
        name = (row.get(name_col) or "").strip()
        if not name:
            continue
        columns.append({
            "name": name,
            "type": _normalize_type(row.get(type_col or "", "") or "STRING"),
            "sample": (row.get(sample_col or "", "") or "")[:100],
            "nullable": str(row.get(nullable_col or "", "yes")).upper() not in ("NO", "NOT NULL", "FALSE", "0"),
        })
    return {"tables": [{"name": "source", "columns": columns}]}


# ─────────────────────────────────────────────────────────────────────────────
# BQ CRAWLER — reads INFORMATION_SCHEMA
# ─────────────────────────────────────────────────────────────────────────────

def crawl_bq(project: str, dataset: str, gcp_creds: str = "", target_tables: List[str] = None) -> List[Dict]:
    """Return list of {table, columns:[{name, type, mode}]} from BQ INFORMATION_SCHEMA."""
    import google.auth
    from google.cloud import bigquery
    from google.oauth2 import service_account

    if gcp_creds and os.path.exists(gcp_creds):
        creds = service_account.Credentials.from_service_account_file(
            gcp_creds,
            scopes=["https://www.googleapis.com/auth/bigquery.readonly"],
        )
        client = bigquery.Client(project=project, credentials=creds)
    else:
        client = bigquery.Client(project=project)

    filter_clause = ""
    if target_tables:
        quoted = ", ".join(f"'{t}'" for t in target_tables)
        filter_clause = f"AND c.table_name IN ({quoted})"

    query = f"""
        SELECT
            c.table_name,
            c.column_name,
            c.data_type,
            c.is_nullable
        FROM `{project}.{dataset}.INFORMATION_SCHEMA.COLUMNS` c
        JOIN `{project}.{dataset}.INFORMATION_SCHEMA.TABLES` t
            ON c.table_name = t.table_name
        WHERE t.table_type = 'BASE TABLE' {filter_clause}
        ORDER BY c.table_name, c.ordinal_position
    """
    rows = list(client.query(query).result())

    tables: Dict[str, List] = {}
    for r in rows:
        tbl = r["table_name"]
        if tbl not in tables:
            tables[tbl] = []
        tables[tbl].append({
            "name": r["column_name"],
            "type": r["data_type"],
            "nullable": r["is_nullable"] == "YES",
        })

    return [{"table": t, "columns": cols} for t, cols in tables.items()]


# ─────────────────────────────────────────────────────────────────────────────
# SOURCE DB CRAWLER — multi-dialect
# ─────────────────────────────────────────────────────────────────────────────

_DB_CONN_EXAMPLES = {
    "postgres":   "postgresql+psycopg2://user:password@host:5432/dbname",
    "mysql":      "mysql+pymysql://user:password@host:3306/dbname",
    "mssql":      "mssql+pyodbc://user:password@host:1433/dbname?driver=ODBC+Driver+17+for+SQL+Server",
    "azuresql":   "mssql+pyodbc://user:password@server.database.windows.net:1433/dbname?driver=ODBC+Driver+17+for+SQL+Server",
    "snowflake":  "snowflake://user:password@account/dbname/schema",
    "databricks": "databricks://token:dapi_token@host/dbname",
    "oracle":     "oracle+cx_oracle://user:password@host:1521/?service_name=ORCLCDB",
    "redshift":   "redshift+psycopg2://user:password@host:5439/dbname",
    "teradata":   "teradatasql://user:password@host/dbname",
}

_DB_INSTALL_HINTS = {
    "postgres":   "pip install psycopg2-binary",
    "mysql":      "pip install pymysql",
    "mssql":      "pip install pyodbc",
    "azuresql":   "pip install pyodbc",
    "snowflake":  "pip install snowflake-connector-python",
    "databricks": "pip install databricks-sql-connector",
    "oracle":     "pip install cx_Oracle  # or oracledb",
    "redshift":   "pip install psycopg2-binary",
    "teradata":   "pip install teradatasql",
}

_INFOSYS_QUERY = """
    SELECT table_name, column_name, data_type, is_nullable
    FROM information_schema.columns
    WHERE table_schema NOT IN ('information_schema','pg_catalog','sys','INFORMATION_SCHEMA','performance_schema')
    ORDER BY table_name, ordinal_position
"""


def crawl_source_db(db_type: str, conn_str: str, schema_filter: str = "", table_filter: str = "") -> Dict:
    """Crawl a source database and return schema in parse_schema_file format."""
    db_type = db_type.lower()

    if db_type == "oracle":
        try:
            import cx_Oracle  # type: ignore
        except ImportError:
            try:
                import oracledb as cx_Oracle  # type: ignore
            except ImportError:
                raise RuntimeError(f"Driver for oracle not installed. Run: {_DB_INSTALL_HINTS['oracle']}")
        import sqlalchemy
        try:
            engine = sqlalchemy.create_engine(conn_str)
        except Exception as e:
            raise RuntimeError(f"Oracle connection failed: {e}")
        query = "SELECT table_name, column_name, data_type, nullable FROM all_tab_columns ORDER BY table_name, column_id"
        with engine.connect() as conn:
            rows = conn.execute(sqlalchemy.text(query)).fetchall()
        tables: Dict[str, List] = {}
        for r in rows:
            tbl = str(r[0])
            if tbl not in tables:
                tables[tbl] = []
            tables[tbl].append({
                "name": str(r[1]),
                "type": _normalize_type(str(r[2])),
                "sample": "",
                "nullable": str(r[3]).upper() == "Y",
            })

    elif db_type == "snowflake":
        try:
            import snowflake.connector  # type: ignore
            # Parse conn_str manually for snowflake connector
            import sqlalchemy
            engine = sqlalchemy.create_engine(conn_str)
            with engine.connect() as conn:
                rows = conn.execute(sqlalchemy.text(_INFOSYS_QUERY)).fetchall()
        except ImportError:
            raise RuntimeError(f"Driver for snowflake not installed. Run: {_DB_INSTALL_HINTS['snowflake']}")
        except Exception as e:
            raise RuntimeError(f"Snowflake connection failed: {e}")
        tables = {}
        for r in rows:
            tbl = str(r[0])
            if tbl not in tables:
                tables[tbl] = []
            tables[tbl].append({
                "name": str(r[1]),
                "type": _normalize_type(str(r[2])),
                "sample": "",
                "nullable": str(r[3]).upper() == "YES",
            })

    elif db_type == "databricks":
        try:
            from databricks import sql as dbsql  # type: ignore
        except ImportError:
            raise RuntimeError(f"Driver for databricks not installed. Run: {_DB_INSTALL_HINTS['databricks']}")
        # For databricks we expect conn_str as: token@host/http_path
        # Use sqlalchemy fallback
        try:
            import sqlalchemy
            engine = sqlalchemy.create_engine(conn_str)
            with engine.connect() as conn:
                rows = conn.execute(sqlalchemy.text("SHOW TABLES")).fetchall()
            tables = {}
            with engine.connect() as conn:
                for row in rows:
                    tbl = str(row[1]) if len(row) > 1 else str(row[0])
                    try:
                        cols = conn.execute(sqlalchemy.text(f"DESCRIBE TABLE {tbl}")).fetchall()
                        tables[tbl] = [{"name": str(c[0]), "type": _normalize_type(str(c[1])), "sample": "", "nullable": True} for c in cols if c[0] and not str(c[0]).startswith("#")]
                    except Exception:
                        pass
        except Exception as e:
            raise RuntimeError(f"Databricks connection failed: {e}")

    else:
        # Postgres, MySQL, MSSQL, Azure SQL, Redshift — use SQLAlchemy + INFORMATION_SCHEMA
        try:
            import sqlalchemy
        except ImportError:
            raise RuntimeError("sqlalchemy not installed. Run: pip install sqlalchemy")

        driver_map = {
            "postgres": "psycopg2",
            "mysql":    "pymysql",
            "mssql":    "pyodbc",
            "azuresql": "pyodbc",
            "redshift": "psycopg2",
        }
        driver = driver_map.get(db_type, "")
        try:
            engine = sqlalchemy.create_engine(conn_str)
        except Exception as e:
            raise RuntimeError(f"Could not create engine for {db_type}: {e}")

        try:
            with engine.connect() as conn:
                rows = conn.execute(sqlalchemy.text(_INFOSYS_QUERY)).fetchall()
        except Exception as e:
            install = _DB_INSTALL_HINTS.get(db_type, "")
            raise RuntimeError(f"Database query failed for {db_type}: {e}. Hint: {install}")

        tables = {}
        for r in rows:
            tbl = str(r[0])
            if tbl not in tables:
                tables[tbl] = []
            tables[tbl].append({
                "name": str(r[1]),
                "type": _normalize_type(str(r[2])),
                "sample": "",
                "nullable": str(r[3]).upper() == "YES",
            })

    # Apply table filter
    if table_filter:
        allowed = {t.strip().lower() for t in table_filter.split(",") if t.strip()}
        tables = {k: v for k, v in tables.items() if k.lower() in allowed}

    # Apply schema filter (for drivers that don't support it natively in INFORMATION_SCHEMA)
    if schema_filter:
        schemas = {s.strip().lower() for s in schema_filter.split(",") if s.strip()}
        # If the table names are schema-qualified, filter; otherwise skip
        filtered = {}
        for tbl, cols in tables.items():
            parts = tbl.split(".")
            if len(parts) == 2 and parts[0].lower() not in schemas:
                continue
            filtered[tbl] = cols
        if filtered:
            tables = filtered

    result_tables = [{"name": tbl, "columns": cols} for tbl, cols in tables.items()]
    return {"tables": result_tables}


# ─────────────────────────────────────────────────────────────────────────────
# CONFIDENCE SCORING — deterministic floor before LLM
# ─────────────────────────────────────────────────────────────────────────────

def _name_score(src: str, tgt: str) -> float:
    try:
        from rapidfuzz import fuzz
        s1 = src.lower().replace("_", "").replace("-", "")
        s2 = tgt.lower().replace("_", "").replace("-", "")
        return fuzz.ratio(s1, s2) / 100.0
    except ImportError:
        # Fallback: common-chars ratio
        s1, s2 = src.lower(), tgt.lower()
        common = sum(c in s2 for c in s1)
        return common / max(len(s1), len(s2), 1)


_TYPE_COMPAT = {
    ("STRING", "STRING"): 1.0, ("INT64", "INT64"): 1.0,
    ("FLOAT64", "FLOAT64"): 1.0, ("BOOLEAN", "BOOLEAN"): 1.0,
    ("DATE", "DATE"): 1.0, ("TIMESTAMP", "TIMESTAMP"): 1.0,
    ("INT64", "FLOAT64"): 0.8, ("FLOAT64", "INT64"): 0.7,
    ("INT64", "STRING"): 0.5, ("STRING", "INT64"): 0.4,
    ("DATE", "TIMESTAMP"): 0.9, ("TIMESTAMP", "DATE"): 0.8,
    ("STRING", "DATE"): 0.5, ("STRING", "TIMESTAMP"): 0.5,
    ("NUMERIC", "FLOAT64"): 0.9, ("NUMERIC", "INT64"): 0.8,
}


def _type_score(src_type: str, tgt_type: str) -> float:
    key = (_normalize_type(src_type), _normalize_type(tgt_type))
    return _TYPE_COMPAT.get(key, 0.3)


def compute_confidence(name_sim: float, type_sim: float, llm_score: float) -> float:
    """Weighted composite: 30% name + 20% type + 50% LLM semantic."""
    return round(name_sim * 0.30 + type_sim * 0.20 + llm_score * 0.50, 3)


def conf_tier(score: float) -> str:
    if score >= 0.80:
        return "high"
    if score >= 0.50:
        return "medium"
    if score > 0:
        return "low"
    return "none"


# ─────────────────────────────────────────────────────────────────────────────
# AGENTIC MAPPING PIPELINE  (runs in background asyncio task)
# ─────────────────────────────────────────────────────────────────────────────

MAPPING_SYSTEM = """You are a senior data engineer performing source-to-target column mapping.
For each source column, pick the BEST matching target column from the candidate tables.
Return ONLY valid JSON — no markdown, no prose.

Output format:
[
  {
    "src_field": "<source column name>",
    "tgt_table": "<target table>",
    "tgt_column": "<target column>",
    "mapping_type": "<Direct|Derived|Lookup|Constant|Expression|Unused>",
    "business_logic": "<transformation expression or description, or null>",
    "llm_confidence": <0.0-1.0 float>,
    "rationale": "<one-line explanation>"
  },
  ...
]

Rules:
- mapping_type=Unused when there is truly no sensible target.
- llm_confidence reflects your certainty (1.0=certain, 0.0=no match).
- business_logic: use BQ SQL where possible (e.g. CAST(x AS INT64), DATE(x), UPPER(TRIM(x))).
- Never fabricate target columns — only use columns listed in the target schema.
"""

SQL_SYSTEM = """You are a BigQuery SQL expert. Given the approved source-to-target mapping,
generate a single production-quality CREATE OR REPLACE TABLE statement.
Rules:
- Use `project.dataset.table` backtick format for all table references.
- Sanitize all column/table names: snake_case, remove spaces, avoid BQ reserved words.
- Include audit columns: _loaded_at TIMESTAMP, _source_run_id STRING, _row_rank INT64 (ROW_NUMBER OVER PARTITION BY primary key).
- Apply business_logic expressions exactly as specified.
- Add explanatory SQL comments for derived/lookup mappings.
- Return ONLY the SQL, no markdown fences.
"""


def _build_mapping_system(session: Dict) -> str:
    """Build context-aware system prompt for L3 mapping stage."""
    base = MAPPING_SYSTEM
    instructions = session.get("instructions", "")
    jira_ctx = session.get("jira_context", {})
    memory = session.get("mapping_memory", [])

    extras = []
    if instructions:
        extras.append(f"USER INSTRUCTIONS (follow strictly):\n{instructions}")
    if jira_ctx.get("summary"):
        extras.append(f"BUSINESS CONTEXT (from Jira):\n{jira_ctx['summary']}")
    if memory:
        mem_text = "\n".join(
            f"  {m['src']} -> {m['tgt']} [{m['type']}]: {m['logic']}"
            for m in memory[-20:]
        )
        extras.append(f"APPROVED MAPPING PATTERNS (learn from these):\n{mem_text}")

    if extras:
        return base + "\n\n" + "\n\n".join(extras)
    return base


async def _emit(session_id: str, event: str, data: Any):
    q = _sse_queues.get(session_id)
    if q:
        await q.put({"event": event, "data": data})


async def _run_pipeline(session_id: str):
    session = _sessions[session_id]
    llm = _make_llm(session)

    async def emit(event: str, data: Any):
        await _emit(session_id, event, data)
        # Mirror to session log
        session.setdefault("log", []).append({"ts": _now(), "event": event, "data": data})

    try:
        # ── L1: Parse uploaded schema ────────────────────────────────────────
        session["stage"] = "L1"
        await emit("stage", {"stage": "L1", "status": "running", "msg": "Parsing source schema…"})
        await asyncio.sleep(0.1)

        schema_data = session.get("schema_data")
        if not schema_data:
            raise RuntimeError("No schema uploaded. Upload an XLSX or CSV file first.")
        src_tables = schema_data.get("tables", [])
        if not src_tables:
            raise RuntimeError("Schema file parsed 0 columns. Check the file format.")

        total_cols = sum(len(t["columns"]) for t in src_tables)
        await emit("stage", {"stage": "L1", "status": "done",
                              "msg": f"Parsed {len(src_tables)} table(s) · {total_cols} source columns"})
        session["l1_done"] = True

        # ── L2: Crawl target schema (BQ or custom files) ──────────────────────
        session["stage"] = "L2"

        if session.get("target_mode") == "files" and session.get("target_files_data"):
            # Use custom uploaded target files
            await emit("stage", {"stage": "L2", "status": "running", "msg": "Loading custom target files…"})
            bq_tables = session["target_files_data"]
            total_tgt_cols = sum(len(t["columns"]) for t in bq_tables)
            await emit("stage", {"stage": "L2", "status": "done",
                                  "msg": f"Using custom target files — {len(bq_tables)} table(s) · {total_tgt_cols} columns"})
        else:
            await emit("stage", {"stage": "L2", "status": "running", "msg": "Crawling BigQuery INFORMATION_SCHEMA…"})

            cfg = session.get("bq_config", {})
            project   = cfg.get("project") or _BQ_PROJECT
            dataset   = cfg.get("dataset") or _BQ_DATASET
            gcp_creds = cfg.get("gcp_creds") or _GCP_CREDS
            tgt_filter = [t.strip() for t in cfg.get("target_tables", "").split(",") if t.strip()]

            if not project or not dataset:
                raise RuntimeError("BQ Project ID and Dataset are required. Configure them in the BQ panel.")

            try:
                bq_tables = await asyncio.to_thread(crawl_bq, project, dataset, gcp_creds, tgt_filter or None)
            except Exception as e:
                raise RuntimeError(f"BigQuery crawl failed: {e}")

            if not bq_tables:
                raise RuntimeError(f"No tables found in {project}.{dataset}. Check project/dataset and permissions.")

            total_tgt_cols = sum(len(t["columns"]) for t in bq_tables)
            await emit("stage", {"stage": "L2", "status": "done",
                                  "msg": f"Crawled {len(bq_tables)} BQ tables · {total_tgt_cols} target columns"})

        session["bq_tables"] = bq_tables
        session["l2_done"] = True

        # Gate 1: auto-approved (no shortlist gate in this flow)
        await emit("gate", {"gate": "gate1", "status": "auto_approved",
                             "msg": "Gate 1 auto-approved — proceeding to semantic mapping"})

        # ── L3: Agentic Semantic Mapping (batched per source table) ───────────
        session["stage"] = "L3"
        await emit("stage", {"stage": "L3", "status": "running", "msg": "Starting semantic mapping…"})

        # Build target schema summary for the prompt
        tgt_summary = "\n".join(
            f"Table: {t['table']}\n  Columns: " +
            ", ".join(f"{c['name']}({c['type']})" for c in t["columns"])
            for t in bq_tables
        )

        # Build context-aware system prompt
        mapping_system = _build_mapping_system(session)

        all_mappings: List[Dict] = []
        processed = 0

        for src_table in src_tables:
            cols = src_table["columns"]
            tbl_name = src_table["name"]
            BATCH = 15

            for i in range(0, len(cols), BATCH):
                batch = cols[i: i + BATCH]
                src_desc = "\n".join(
                    f"  - {c['name']} ({c['type']})" + (f" sample={c['sample']}" if c.get("sample") else "")
                    for c in batch
                )

                prompt = (
                    f"SOURCE TABLE: {tbl_name}\n"
                    f"SOURCE COLUMNS TO MAP:\n{src_desc}\n\n"
                    f"TARGET SCHEMA:\n{tgt_summary}\n\n"
                    "Map each source column to its best target. Return JSON array."
                )

                try:
                    raw_result = await asyncio.to_thread(llm.complete_json, mapping_system, prompt)
                    if isinstance(raw_result, dict):
                        raw_result = raw_result.get("mappings", [raw_result])
                except Exception as e:
                    logger.warning("LLM mapping batch failed: %s", e)
                    raw_result = []

                # Enrich with deterministic confidence floor
                for item in raw_result:
                    src_col = next((c for c in batch if c["name"] == item.get("src_field")), None)
                    tgt_table_name = item.get("tgt_table", "")
                    tgt_col_name   = item.get("tgt_column", "")

                    # Find target column type
                    tgt_col_type = "STRING"
                    for tbl in bq_tables:
                        if tbl["table"] == tgt_table_name:
                            for col in tbl["columns"]:
                                if col["name"] == tgt_col_name:
                                    tgt_col_type = col["type"]
                                    break

                    name_sim = _name_score(item.get("src_field", ""), tgt_col_name) if tgt_col_name else 0.0
                    type_sim = _type_score(src_col["type"] if src_col else "STRING", tgt_col_type) if tgt_col_name else 0.0
                    llm_conf  = float(item.get("llm_confidence", 0.5))

                    is_unused = item.get("mapping_type", "").lower() == "unused" or not tgt_col_name
                    confidence = 0.0 if is_unused else compute_confidence(name_sim, type_sim, llm_conf)

                    row_id = str(uuid.uuid4())
                    all_mappings.append({
                        "id":             row_id,
                        "src_table":      tbl_name,
                        "src_field":      item.get("src_field", ""),
                        "src_type":       src_col["type"] if src_col else "STRING",
                        "tgt_table":      tgt_table_name if not is_unused else "",
                        "tgt_column":     tgt_col_name if not is_unused else "",
                        "tgt_type":       tgt_col_type if not is_unused else "",
                        "mapping_type":   item.get("mapping_type", "Direct"),
                        "business_logic": item.get("business_logic", "") or "",
                        "confidence":     confidence,
                        "tier":           conf_tier(confidence),
                        "status":         "unmapped" if is_unused else ("review" if confidence < 0.8 else "mapped"),
                        "rationale":      item.get("rationale", ""),
                        "llm_confidence": llm_conf,
                        "name_sim":       round(name_sim, 3),
                        "type_sim":       round(type_sim, 3),
                        "modified":       False,
                    })

                processed += len(batch)
                total = sum(len(t["columns"]) for t in src_tables)
                await emit("progress", {"processed": processed, "total": total,
                                         "msg": f"Mapped {processed}/{total} columns…"})

        session["mappings"] = all_mappings
        n_mapped   = sum(1 for m in all_mappings if m["status"] == "mapped")
        n_review   = sum(1 for m in all_mappings if m["status"] == "review")
        n_unmapped = sum(1 for m in all_mappings if m["status"] == "unmapped")
        avg_conf   = sum(m["confidence"] for m in all_mappings) / max(len(all_mappings), 1)

        session["stats"] = {
            "total": len(all_mappings),
            "mapped": n_mapped,
            "review": n_review,
            "unmapped": n_unmapped,
            "avg_confidence": round(avg_conf, 3),
        }

        await emit("stage", {
            "stage": "L3", "status": "done",
            "msg": f"Mapping complete · {n_mapped} auto-mapped · {n_review} need review · {n_unmapped} unmapped",
            "stats": session["stats"],
        })
        session["l3_done"] = True

        # Gate 2: requires human review — pause here
        session["stage"] = "gate2"
        await emit("gate", {"gate": "gate2", "status": "awaiting",
                             "msg": "Gate 2: Review and edit the mapping table, then click 'Approve & Generate SQL'"})

        session["status"] = "review"
        await emit("status", {"status": "review", "msg": "Ready for human review"})

    except Exception as e:
        logger.exception("Pipeline error: %s", e)
        session["status"] = "error"
        session["error"]  = str(e)
        await emit("error", {"msg": str(e)})
    finally:
        session["running"] = False
        q = _sse_queues.get(session_id)
        if q:
            await q.put(None)  # sentinel


async def _run_sql_generation(session_id: str):
    session = _sessions[session_id]
    llm     = _make_llm(session)

    async def emit(event: str, data: Any):
        await _emit(session_id, event, data)

    try:
        session["stage"] = "L4"
        await emit("stage", {"stage": "L4", "status": "running", "msg": "Generating materialized BigQuery SQL…"})

        mappings = session.get("mappings", [])
        cfg      = session.get("bq_config", {})
        project  = cfg.get("project") or _BQ_PROJECT
        dataset  = cfg.get("dataset") or _BQ_DATASET

        mapped_rows = [m for m in mappings if m["status"] != "unmapped" and m.get("tgt_table")]
        if not mapped_rows:
            raise RuntimeError("No mapped rows to generate SQL from.")

        # Group by target table
        tgt_groups: Dict[str, List] = {}
        for m in mapped_rows:
            tgt_groups.setdefault(m["tgt_table"], []).append(m)

        sql_blocks = []
        for tgt_table, rows in tgt_groups.items():
            src_tables = list({r["src_table"] for r in rows})
            primary_src = src_tables[0]

            mapping_desc = "\n".join(
                f"  {r['src_field']} ({r['src_type']}) -> {r['tgt_column']} ({r['tgt_type']}) "
                f"[{r['mapping_type']}] logic: {r['business_logic'] or 'direct'}"
                for r in rows
            )

            prompt = (
                f"BQ Project: {project}\n"
                f"BQ Dataset: {dataset}\n"
                f"Source table: {primary_src}\n"
                f"Target table: {tgt_table}\n\n"
                f"Approved mappings:\n{mapping_desc}\n\n"
                "Generate a single CREATE OR REPLACE TABLE SQL statement."
            )

            sql = await asyncio.to_thread(llm.complete, SQL_SYSTEM, prompt, 0.05, 4096)
            # Strip any markdown fences
            sql = re.sub(r"```(?:sql)?\s*", "", sql).replace("```", "").strip()
            sql_blocks.append(f"-- ═══ Target: {project}.{dataset}.{tgt_table} ═══\n{sql}\n")

        final_sql = (
            f"-- Auto-generated by DataMapper · {_now()}\n"
            f"-- Session: {session_id[:8]}\n"
            f"-- Mapped: {len(mapped_rows)} columns · Avg confidence: {session.get('stats', {}).get('avg_confidence', 0):.0%}\n\n"
        ) + "\n".join(sql_blocks)

        session["generated_sql"] = final_sql
        session["status"] = "done"
        session["stage"]  = "done"

        await emit("stage", {"stage": "L4", "status": "done", "msg": "SQL generated successfully"})
        await emit("status", {"status": "done", "msg": "Pipeline complete"})

    except Exception as e:
        logger.exception("SQL gen error: %s", e)
        await emit("error", {"msg": str(e)})
    finally:
        session["running"] = False
        q = _sse_queues.get(session_id)
        if q:
            await q.put(None)


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _session_or_404(session_id: str) -> Dict:
    if session_id not in _sessions:
        raise HTTPException(404, f"Session {session_id!r} not found")
    return _sessions[session_id]


# ─────────────────────────────────────────────────────────────────────────────
# FASTAPI APP
# ─────────────────────────────────────────────────────────────────────────────

app = FastAPI(title="DataMapper", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve frontend
_STATIC = Path(__file__).parent
if (_STATIC / "index.html").exists():
    @app.get("/", include_in_schema=False)
    async def root():
        return FileResponse(_STATIC / "index.html")


# ── Session lifecycle ─────────────────────────────────────────────────────────

class SessionCreate(BaseModel):
    name: Optional[str] = None
    instructions: Optional[str] = None


@app.post("/api/sessions")
async def create_session(body: Optional[SessionCreate] = None):
    sid = str(uuid.uuid4())
    _sessions[sid] = {
        "id":                sid,
        "created_at":        _now(),
        "status":            "new",
        "stage":             "idle",
        "running":           False,
        "log":               [],
        "mappings":          [],
        "stats":             {},
        "bq_config":         {},
        "api_config":        {},
        "name":              (body.name if body else None) or f"Session {sid[:6]}",
        "instructions":      (body.instructions if body else None) or "",
        "mapping_memory":    [],
        "jira_context":      {},
        "target_mode":       "bq",
        "target_files_data": None,
    }
    return {"session_id": sid}


@app.get("/api/sessions")
async def list_sessions():
    return [
        {
            "id":         s["id"],
            "status":     s["status"],
            "stage":      s["stage"],
            "created_at": s["created_at"],
            "stats":      s.get("stats", {}),
            "filename":   s.get("filename", ""),
        }
        for s in sorted(_sessions.values(), key=lambda x: x["created_at"], reverse=True)
    ]


@app.get("/api/sessions/{sid}")
async def get_session(sid: str):
    s = _session_or_404(sid)
    return {
        "id":          s["id"],
        "status":      s["status"],
        "stage":       s["stage"],
        "created_at":  s["created_at"],
        "stats":       s.get("stats", {}),
        "bq_config":   {k: v for k, v in s.get("bq_config", {}).items() if k != "gcp_creds"},
        "api_config":  {k: ("***" if "key" in k else v) for k, v in s.get("api_config", {}).items()},
        "filename":    s.get("filename", ""),
        "schema_data": s.get("schema_data"),
        "error":       s.get("error"),
        "log":         s.get("log", [])[-50:],
    }


# ── Schema upload ─────────────────────────────────────────────────────────────

@app.post("/api/sessions/{sid}/upload")
async def upload_schema(sid: str, file: UploadFile = File(...)):
    s = _session_or_404(sid)
    content = await file.read()
    try:
        schema_data = parse_schema_file(content, file.filename)
    except Exception as e:
        raise HTTPException(422, str(e))
    s["schema_data"] = schema_data
    s["filename"]    = file.filename
    s["status"]      = "schema_uploaded"
    total = sum(len(t["columns"]) for t in schema_data["tables"])
    return {
        "ok":     True,
        "tables": len(schema_data["tables"]),
        "columns": total,
        "preview": schema_data["tables"][0]["columns"][:8] if schema_data["tables"] else [],
        "table_names": [t["name"] for t in schema_data["tables"]],
    }


# ── Jira context ──────────────────────────────────────────────────────────────

class JiraContextRequest(BaseModel):
    jira_url:   Optional[str] = ""
    jira_email: Optional[str] = ""
    jira_token: Optional[str] = ""
    ticket_url: Optional[str] = ""


@app.post("/api/sessions/{sid}/jira-context")
async def fetch_jira_context(sid: str, req: JiraContextRequest):
    s = _session_or_404(sid)

    base_url  = req.jira_url  or _JIRA_URL
    email     = req.jira_email or _JIRA_EMAIL
    token     = req.jira_token or _JIRA_TOKEN
    ticket    = req.ticket_url or ""

    if not base_url or not token:
        raise HTTPException(422, "Jira base URL and API token are required (or set JIRA_URL/JIRA_TOKEN in .env)")

    # Extract issue key from ticket URL or treat as key directly
    issue_key = ticket
    if "/" in ticket:
        issue_key = ticket.rstrip("/").split("/")[-1]

    if not issue_key:
        raise HTTPException(422, "Jira ticket URL or issue key is required")

    api_url = f"{base_url.rstrip('/')}/rest/api/3/issue/{issue_key}"
    import base64
    creds = base64.b64encode(f"{email}:{token}".encode()).decode()

    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(api_url, headers={
                "Authorization": f"Basic {creds}",
                "Accept": "application/json",
            })
        if resp.status_code == 401:
            raise HTTPException(401, "Jira authentication failed. Check email and API token.")
        if resp.status_code == 404:
            raise HTTPException(404, f"Jira issue {issue_key!r} not found.")
        resp.raise_for_status()
        issue_data = resp.json()
    except httpx.RequestError as e:
        raise HTTPException(502, f"Could not reach Jira: {e}")

    fields = issue_data.get("fields", {})
    summary_text = fields.get("summary", "")
    desc = fields.get("description", {})
    desc_text = ""
    if isinstance(desc, dict):
        for block in desc.get("content", []):
            for item in block.get("content", []):
                if item.get("type") == "text":
                    desc_text += item.get("text", "") + " "
    elif isinstance(desc, str):
        desc_text = desc

    llm = _make_llm(s)
    prompt = f"""Jira Issue: {issue_key}
Summary: {summary_text}
Description: {desc_text[:2000]}

Extract from this Jira story:
1. Intent summary (1-2 sentences, what data engineering work is needed)
2. Source system hint (what source system/database is mentioned)
3. Target system hint (what target/destination is mentioned)
4. Key business rules (bullet list, max 5)

Return JSON: {{"summary": "...", "source_hint": "...", "target_hint": "...", "business_rules": ["..."]}}"""

    system = "You are a data engineering analyst. Extract key mapping context from a Jira story. Return only valid JSON."
    try:
        ctx = await asyncio.to_thread(llm.complete_json, system, prompt)
    except Exception as e:
        ctx = {"summary": summary_text, "source_hint": "", "target_hint": "", "business_rules": []}

    ctx["issue_key"] = issue_key
    ctx["jira_url"] = f"{base_url.rstrip('/')}/browse/{issue_key}"
    s["jira_context"] = ctx
    return {"ok": True, "context": ctx}


# ── Source DB connector ────────────────────────────────────────────────────────

class SourceConnectRequest(BaseModel):
    db_type:           str
    connection_string: str
    schema_filter:     Optional[str] = ""
    table_filter:      Optional[str] = ""


@app.post("/api/sessions/{sid}/source-connect")
async def source_connect(sid: str, req: SourceConnectRequest):
    s = _session_or_404(sid)
    try:
        schema_data = await asyncio.to_thread(
            crawl_source_db,
            req.db_type,
            req.connection_string,
            req.schema_filter or "",
            req.table_filter or "",
        )
    except RuntimeError as e:
        raise HTTPException(422, str(e))
    except Exception as e:
        raise HTTPException(500, f"Connection failed: {e}")

    s["schema_data"]  = schema_data
    s["source_type"]  = req.db_type
    # Extract host from connection string for display
    import re as _re
    host_match = _re.search(r"@([^/:]+)", req.connection_string)
    host = host_match.group(1) if host_match else "unknown"
    s["source_conn_display"] = f"{req.db_type}://{host}"
    s["filename"]    = s["source_conn_display"]
    s["status"]      = "schema_uploaded"

    total = sum(len(t["columns"]) for t in schema_data["tables"])
    return {
        "ok":          True,
        "tables":      len(schema_data["tables"]),
        "columns":     total,
        "table_names": [t["name"] for t in schema_data["tables"]],
        "preview":     schema_data["tables"][0]["columns"][:8] if schema_data["tables"] else [],
    }


# ── Target files upload ────────────────────────────────────────────────────────

@app.post("/api/sessions/{sid}/target-files")
async def upload_target_files(sid: str, files: List[UploadFile] = File(...)):
    s = _session_or_404(sid)
    target_tables = []

    for f in files:
        content = await f.read()
        table_name = f.filename.rsplit(".", 1)[0] if "." in f.filename else f.filename

        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        fieldnames = [h.lower().strip() for h in (reader.fieldnames or [])]

        name_col     = next((h for h in fieldnames if h in ("column_name", "column", "field", "name")), fieldnames[0] if fieldnames else "column_name")
        type_col     = next((h for h in fieldnames if h in ("data_type", "type", "datatype")), None)
        nullable_col = next((h for h in fieldnames if h in ("is_nullable", "nullable", "null")), None)

        columns = []
        for row in reader:
            col_name = (row.get(name_col) or "").strip()
            if not col_name:
                continue
            columns.append({
                "name":     col_name,
                "type":     _normalize_type(row.get(type_col or "", "") or "STRING") if type_col else "STRING",
                "nullable": str(row.get(nullable_col or "", "YES")).upper() not in ("NO", "NOT NULL", "FALSE", "0") if nullable_col else True,
            })

        if columns:
            target_tables.append({"table": table_name, "columns": columns})

    s["target_mode"]       = "files"
    s["target_files_data"] = target_tables

    total_cols = sum(len(t["columns"]) for t in target_tables)
    return {
        "ok":           True,
        "tables":       len(target_tables),
        "table_names":  [t["table"] for t in target_tables],
        "total_columns": total_cols,
    }


# ── Session summary ────────────────────────────────────────────────────────────

@app.get("/api/sessions/{sid}/summary")
async def get_session_summary(sid: str):
    s = _session_or_404(sid)
    mappings = s.get("mappings", [])
    if not mappings:
        return {"ready": False}

    llm = _make_llm(s)
    stats = s.get("stats", {})
    prompt = f"""Session: {sid[:8]}
Source: {s.get('filename') or s.get('source_type', 'unknown')}
Instructions given: {s.get('instructions', 'None')}
Stats: {json.dumps(stats)}
Top mappings (first 30): {json.dumps(mappings[:30], indent=2)}

Write a structured summary with:
1. Overview (2 sentences)
2. Key mapping decisions made (bullet points, max 8)
3. Fields needing attention (unmapped or low confidence, max 5)
4. Recommended next steps (max 3)
Keep total under 400 words."""

    system = "You are a data engineering expert. Summarize a mapping session concisely."
    try:
        summary_text = await asyncio.to_thread(llm.complete, system, prompt, 0.1, 1024)
    except Exception as e:
        summary_text = f"Summary unavailable: {e}"

    return {"ready": True, "summary": summary_text, "stats": stats}


# ── BQ config ─────────────────────────────────────────────────────────────────

class BQConfig(BaseModel):
    project:       str
    dataset:       str
    region:        Optional[str] = "us-central1"
    gcp_creds:     Optional[str] = ""
    target_tables: Optional[str] = ""


@app.post("/api/sessions/{sid}/bq-config")
async def set_bq_config(sid: str, cfg: BQConfig):
    s = _session_or_404(sid)
    s["bq_config"] = cfg.model_dump()
    return {"ok": True}


@app.post("/api/sessions/{sid}/bq-test")
async def test_bq(sid: str):
    s = _session_or_404(sid)
    cfg = s.get("bq_config", {})
    project = cfg.get("project") or _BQ_PROJECT
    dataset = cfg.get("dataset") or _BQ_DATASET
    if not project or not dataset:
        raise HTTPException(422, "Project and dataset required")
    try:
        tables = await asyncio.to_thread(crawl_bq, project, dataset, cfg.get("gcp_creds") or _GCP_CREDS, None)
        return {"ok": True, "tables": len(tables), "table_names": [t["table"] for t in tables]}
    except Exception as e:
        raise HTTPException(422, str(e))


# ── API key config ─────────────────────────────────────────────────────────────

class APIConfig(BaseModel):
    provider: str = "claude"      # claude | deepseek | custom
    api_key:  Optional[str] = ""
    base_url: Optional[str] = ""
    model:    Optional[str] = ""


@app.post("/api/sessions/{sid}/api-config")
async def set_api_config(sid: str, cfg: APIConfig):
    s = _session_or_404(sid)
    s["api_config"] = cfg.model_dump()
    return {"ok": True}


@app.get("/api/global-config")
async def global_config():
    return {
        "default_provider":  _DEFAULT_PROVIDER,
        "has_anthropic_key": bool(_ANTHROPIC_API_KEY),
        "has_deepseek_key":  bool(_DEEPSEEK_API_KEY),
        "has_bq_project":    bool(_BQ_PROJECT),
        "bq_project":        _BQ_PROJECT,
        "bq_dataset":        _BQ_DATASET,
        "claude_model":      _CLAUDE_MODEL,
        "deepseek_model":    _DEEPSEEK_MODEL,
        "gcp_creds":         bool(_GCP_CREDS),
    }


# ── Pipeline control ──────────────────────────────────────────────────────────

@app.post("/api/sessions/{sid}/run")
async def run_pipeline(sid: str):
    s = _session_or_404(sid)
    if s.get("running"):
        raise HTTPException(409, "Pipeline already running")
    if not s.get("schema_data"):
        raise HTTPException(422, "Upload a schema file first")
    s["status"]  = "running"
    s["running"] = True
    s["error"]   = None
    s["mappings"] = []
    s["stats"]   = {}
    # Create fresh SSE queue
    _sse_queues[sid] = asyncio.Queue()
    asyncio.create_task(_run_pipeline(sid))
    return {"ok": True, "msg": "Pipeline started"}


@app.post("/api/sessions/{sid}/approve-gate2")
async def approve_gate2(sid: str):
    s = _session_or_404(sid)
    if s.get("status") != "review":
        raise HTTPException(409, "Session is not at Gate 2 review stage")
    s["status"]  = "running"
    s["running"] = True
    _sse_queues[sid] = asyncio.Queue()
    asyncio.create_task(_run_sql_generation(sid))
    return {"ok": True, "msg": "SQL generation started"}


# ── SSE event stream ──────────────────────────────────────────────────────────

@app.get("/api/sessions/{sid}/events")
async def sse_stream(sid: str, request: Request):
    s = _session_or_404(sid)

    async def generator() -> AsyncIterator[str]:
        # Send current state immediately
        yield f"data: {json.dumps({'event':'state','data':{'status':s['status'],'stage':s['stage']}})}\n\n"

        q = _sse_queues.get(sid)
        if not q:
            return

        while True:
            if await request.is_disconnected():
                break
            try:
                msg = await asyncio.wait_for(q.get(), timeout=20.0)
            except asyncio.TimeoutError:
                yield ": heartbeat\n\n"
                continue
            if msg is None:
                yield f"data: {json.dumps({'event':'done','data':{}})}\n\n"
                break
            yield f"data: {json.dumps({'event': msg['event'], 'data': msg['data']})}\n\n"

    return StreamingResponse(generator(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ── Mapping table CRUD ─────────────────────────────────────────────────────────

@app.get("/api/sessions/{sid}/mappings")
async def get_mappings(sid: str):
    s = _session_or_404(sid)
    return {"mappings": s.get("mappings", []), "stats": s.get("stats", {})}


class MappingPatch(BaseModel):
    tgt_table:      Optional[str] = None
    tgt_column:     Optional[str] = None
    mapping_type:   Optional[str] = None
    business_logic: Optional[str] = None
    status:         Optional[str] = None


@app.patch("/api/sessions/{sid}/mappings/{row_id}")
async def patch_mapping(sid: str, row_id: str, patch: MappingPatch):
    s = _session_or_404(sid)
    mappings = s.get("mappings", [])
    row = next((m for m in mappings if m["id"] == row_id), None)
    if not row:
        raise HTTPException(404, "Mapping row not found")

    for field, val in patch.model_dump(exclude_none=True).items():
        row[field] = val
    row["modified"] = True

    # Recompute status if tgt was assigned
    if row.get("tgt_column") and row.get("status") == "unmapped":
        row["status"] = "review"
    if row.get("status") == "unmapped" and not row.get("tgt_column"):
        row["confidence"] = 0.0
        row["tier"] = "none"

    # Recompute stats
    n_mapped   = sum(1 for m in mappings if m["status"] == "mapped")
    n_review   = sum(1 for m in mappings if m["status"] == "review")
    n_unmapped = sum(1 for m in mappings if m["status"] == "unmapped")
    avg_conf   = sum(m["confidence"] for m in mappings) / max(len(mappings), 1)
    s["stats"]  = {"total": len(mappings), "mapped": n_mapped,
                   "review": n_review, "unmapped": n_unmapped,
                   "avg_confidence": round(avg_conf, 3)}
    return {"ok": True, "row": row, "stats": s["stats"]}


@app.post("/api/sessions/{sid}/mappings/{row_id}/no-mapping")
async def set_no_mapping(sid: str, row_id: str):
    s = _session_or_404(sid)
    row = next((m for m in s.get("mappings", []) if m["id"] == row_id), None)
    if not row:
        raise HTTPException(404, "Row not found")
    row.update({"tgt_table": "", "tgt_column": "", "status": "unmapped",
                "confidence": 0.0, "tier": "none", "modified": True})
    return {"ok": True}


@app.post("/api/sessions/{sid}/mappings/{row_id}/approve")
async def approve_mapping(sid: str, row_id: str):
    s = _session_or_404(sid)
    row = next((m for m in s.get("mappings", []) if m["id"] == row_id), None)
    if not row:
        raise HTTPException(404, "Row not found")
    row["status"] = "mapped"
    row["modified"] = True
    # Record to session memory for future runs
    s.setdefault("mapping_memory", []).append({
        "src": row.get("src_field", ""),
        "tgt": row.get("tgt_column", ""),
        "type": row.get("mapping_type", ""),
        "logic": row.get("business_logic", ""),
    })
    return {"ok": True}


# ── Export ────────────────────────────────────────────────────────────────────

@app.get("/api/sessions/{sid}/export/csv")
async def export_csv(sid: str):
    s = _session_or_404(sid)
    mappings = s.get("mappings", [])
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=[
        "src_table", "src_field", "src_type",
        "tgt_table", "tgt_column", "tgt_type",
        "mapping_type", "business_logic",
        "confidence", "tier", "status", "rationale",
    ])
    w.writeheader()
    for m in mappings:
        w.writerow({k: m.get(k, "") for k in w.fieldnames})
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="stm_{sid[:8]}.csv"'},
    )


@app.get("/api/sessions/{sid}/export/xlsx")
async def export_xlsx(sid: str):
    s = _session_or_404(sid)
    mappings = s.get("mappings", [])

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "STM Mappings"

    HEADERS = [
        "Src Table", "Src Field", "Src Type", "→",
        "Tgt Table", "Tgt Column", "Tgt Type",
        "Mapping Type", "Business Logic",
        "Confidence %", "Tier", "Status", "Rationale",
    ]
    HDR_FILL = PatternFill("solid", fgColor="141413")
    HDR_FONT = Font(bold=True, color="FAFAFA", name="Calibri", size=10)
    TIER_FILL = {
        "high":   PatternFill("solid", fgColor="EEF4E8"),
        "medium": PatternFill("solid", fgColor="FEF6EC"),
        "low":    PatternFill("solid", fgColor="FDF0F0"),
        "none":   PatternFill("solid", fgColor="F2F0E8"),
    }

    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, m in enumerate(mappings, 2):
        values = [
            m.get("src_table", ""),  m.get("src_field", ""),  m.get("src_type", ""),
            "→" if m.get("tgt_column") else "✕",
            m.get("tgt_table", ""),  m.get("tgt_column", ""), m.get("tgt_type", ""),
            m.get("mapping_type", ""),
            m.get("business_logic", ""),
            round((m.get("confidence") or 0) * 100),
            m.get("tier", ""),  m.get("status", ""),  m.get("rationale", ""),
        ]
        fill = TIER_FILL.get(m.get("tier", ""), PatternFill("solid", fgColor="FFFFFF"))
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill

    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col) + 4
        ws.column_dimensions[col[0].column_letter].width = min(w, 52)
    ws.freeze_panes = "A2"

    # ── Summary sheet ──
    ws2 = wb.create_sheet("Summary")
    stats = s.get("stats", {})
    ws2["A1"] = "DataMapper AI — Session Summary"
    ws2["A1"].font = Font(bold=True, size=14, name="Calibri")
    rows2 = [
        ("Session ID",     sid),
        ("Generated At",   _now()),
        ("Source File",    s.get("filename", "")),
        ("",               ""),
        ("Total Fields",   stats.get("total", 0)),
        ("Auto-Mapped",    stats.get("mapped", 0)),
        ("Needs Review",   stats.get("review", 0)),
        ("Unmapped",       stats.get("unmapped", 0)),
        ("Avg Confidence", f"{round((stats.get('avg_confidence') or 0) * 100)}%"),
    ]
    for i, (k, v) in enumerate(rows2, 3):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True, name="Calibri")
        ws2.cell(row=i, column=2, value=v)
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="stm_{sid[:8]}.xlsx"'},
    )


@app.get("/api/sessions/{sid}/export/sql")
async def export_sql(sid: str):
    s = _session_or_404(sid)
    sql = s.get("generated_sql")
    if not sql:
        raise HTTPException(422, "SQL not yet generated. Complete the pipeline first.")
    return StreamingResponse(
        io.BytesIO(sql.encode()),
        media_type="text/plain",
        headers={"Content-Disposition": f'attachment; filename="stm_{sid[:8]}.sql"'},
    )


@app.get("/api/sessions/{sid}/sql")
async def get_sql(sid: str):
    s = _session_or_404(sid)
    return {"sql": s.get("generated_sql", ""), "ready": bool(s.get("generated_sql"))}


# ── Health ────────────────────────────────────────────────────────────────────

@app.get("/api/health")
async def health():
    return {
        "ok":      True,
        "version": "1.0.0",
        "sessions": len(_sessions),
        "providers": {
            "claude":   bool(_ANTHROPIC_API_KEY),
            "deepseek": bool(_DEEPSEEK_API_KEY),
        },
        "env_loaded": _ENV_PATH.exists(),
    }


@app.get("/api/version")
async def version():
    return {"version": "1.0.0", "name": "DataMapper AI", "stage": "v1"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("server:app", host="0.0.0.0", port=7788, reload=True, log_level="info")
