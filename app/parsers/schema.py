"""
app/parsers/schema.py — parse_schema_file + all Excel/CSV helpers
"""
from __future__ import annotations

import csv
import io
import re
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Optional


def _normalize_type(raw: str) -> str:
    raw = str(raw or "").upper().strip().split("(")[0]
    MAP = {
        "INT": "INT64", "INTEGER": "INT64", "BIGINT": "INT64", "SMALLINT": "INT64",
        "TINYINT": "INT64", "MEDIUMINT": "INT64", "SERIAL": "INT64",
        "FLOAT": "FLOAT64", "DOUBLE": "FLOAT64", "REAL": "FLOAT64",
        "DECIMAL": "NUMERIC", "NUMBER": "NUMERIC", "MONEY": "NUMERIC",
        "BOOL": "BOOLEAN", "BIT": "BOOLEAN",
        "VARCHAR": "STRING", "NVARCHAR": "STRING", "CHAR": "STRING",
        "NCHAR": "STRING", "TEXT": "STRING", "CLOB": "STRING", "NTEXT": "STRING",
        "DATE": "DATE", "DATETIME": "TIMESTAMP", "TIMESTAMP": "TIMESTAMP",
        "DATETIME2": "TIMESTAMP", "SMALLDATETIME": "TIMESTAMP",
        "BLOB": "BYTES", "BINARY": "BYTES", "VARBINARY": "BYTES", "BYTEA": "BYTES",
        "JSON": "JSON", "JSONB": "JSON", "ARRAY": "ARRAY",
        "UUID": "STRING", "UNIQUEIDENTIFIER": "STRING",
    }
    return MAP.get(raw, raw or "STRING")


def _infer_type_from_val(val) -> str:
    """Infer BigQuery type from a Python cell value."""
    from datetime import datetime as _dt, date as _date
    if val is None or val == "":
        return "STRING"
    if isinstance(val, bool):
        return "BOOLEAN"
    if isinstance(val, int):
        return "INT64"
    if isinstance(val, float):
        return "FLOAT64"
    if isinstance(val, _dt):
        return "DATE" if (val.hour == 0 and val.minute == 0 and val.second == 0) else "TIMESTAMP"
    if isinstance(val, _date):
        return "DATE"
    s = str(val).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):                    return "DATE"
    if re.match(r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}", s):      return "TIMESTAMP"
    if s.upper() in ("TRUE", "FALSE", "YES", "NO"):              return "BOOLEAN"
    try:    int(s);   return "INT64"
    except ValueError: pass
    try:    float(s); return "FLOAT64"
    except ValueError: pass
    return "STRING"


def _col_type_from_samples(samples: List) -> str:
    """Pick the most specific type that covers all non-null sample values."""
    types = [_infer_type_from_val(v) for v in samples if v is not None and v != ""]
    if not types:
        return "STRING"
    if "STRING" in types:
        return "STRING"
    if "FLOAT64" in types and "INT64" in types:
        return "FLOAT64"
    return Counter(types).most_common(1)[0][0]


_INFO_SCHEMA_KEYS  = {"table_name", "columns"}
_SCHEMA_DUMP_KEYS  = {"column_name", "field", "column", "name", "field_name"}
_TYPE_KEYS         = {"type", "data_type", "datatype", "dtype", "column_type"}
_NULLABLE_KEYS     = {"nullable", "null", "is_nullable", "required"}
_SAMPLE_KEYS       = {"sample", "example", "sample_value"}


def _parse_sheet(sheet_name: str, rows: List[tuple]) -> Optional[Any]:
    """Parse a single XLSX sheet into a table dict, auto-detecting format."""
    if not rows or not rows[0]:
        return None
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    hdr_lower = [h.lower() for h in header]
    hdr_set   = set(hdr_lower)

    # Format A: Information-schema dump
    if _INFO_SCHEMA_KEYS.issubset(hdr_set):
        tbl_idx  = hdr_lower.index("table_name")
        cols_idx = hdr_lower.index("columns")
        tables: Dict[str, List] = {}
        for row in rows[1:]:
            tname = str(row[tbl_idx]).strip() if row[tbl_idx] else ""
            cstr  = str(row[cols_idx]).strip() if row[cols_idx] else ""
            if not tname or not cstr or tname.lower() in ("none", "nan"):
                continue
            col_names = [c.strip() for c in cstr.replace(";", ",").split(",") if c.strip()]
            if tname not in tables:
                tables[tname] = []
            for cn in col_names:
                tables[tname].append({"name": cn, "type": "STRING", "sample": "", "nullable": True})
        return [{"name": t, "columns": c} for t, c in tables.items()] if tables else None

    # Format B: Schema dump
    name_idx = next((hdr_lower.index(k) for k in ("column_name", "field_name", "field", "column", "name") if k in hdr_set), None)
    type_idx = next((hdr_lower.index(k) for k in _TYPE_KEYS if k in hdr_set), None)
    null_idx = next((hdr_lower.index(k) for k in _NULLABLE_KEYS if k in hdr_set), None)
    samp_idx = next((hdr_lower.index(k) for k in _SAMPLE_KEYS if k in hdr_set), None)

    if name_idx is not None:
        has_types = type_idx is not None and any(
            str(r[type_idx] or "").upper().strip() in (
                "STRING", "INT64", "INT", "VARCHAR", "TEXT", "INTEGER", "FLOAT",
                "FLOAT64", "BOOLEAN", "TIMESTAMP", "DATE", "NUMERIC", "NUMBER",
                "BIGINT", "DOUBLE", "CHAR", "NVARCHAR",
            )
            for r in rows[1:5] if len(r) > type_idx
        )
        first_names = [str(r[name_idx] or "").strip() for r in rows[1:6] if r and r[name_idx]]
        looks_like_cols = all(re.match(r"^[A-Za-z_][A-Za-z0-9_ ]*$", n) for n in first_names if n)

        if has_types or (type_idx is None and looks_like_cols and len(rows) < 500):
            columns = []
            for row in rows[1:]:
                if len(row) <= name_idx:
                    continue
                cname = str(row[name_idx] or "").strip()
                if not cname or cname.lower() in ("none", "nan", ""):
                    continue
                raw_type = str(row[type_idx] or "").strip() if type_idx is not None and len(row) > type_idx else "STRING"
                columns.append({
                    "name": cname,
                    "type": _normalize_type(raw_type),
                    "sample": str(row[samp_idx] or "")[:100] if samp_idx is not None and len(row) > samp_idx else "",
                    "nullable": str(row[null_idx] or "YES").upper() not in ("NO", "NOT NULL", "FALSE", "0") if null_idx is not None and len(row) > null_idx else True,
                })
            if columns:
                return {"name": sheet_name, "columns": columns}

    # Format C: Data file
    valid_cols = [h for h in header if h and re.match(r"^[A-Za-z_\-][A-Za-z0-9_ \-]*$", h)]
    if len(valid_cols) < max(1, len(header) * 0.5):
        return None

    sample_rows = rows[1:11]
    columns = []
    for ci, col_name in enumerate(header):
        if not col_name:
            continue
        samples = [r[ci] if ci < len(r) else None for r in sample_rows]
        col_type = _col_type_from_samples(samples)
        sample_str = str(samples[0])[:100] if samples and samples[0] is not None else ""
        columns.append({"name": col_name, "type": col_type, "sample": sample_str, "nullable": True})

    return {"name": sheet_name, "columns": columns} if columns else None


def _parse_excel(content: bytes) -> Dict[str, Any]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    tables: List[Dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        result = _parse_sheet(sheet_name, rows)
        if result is None:
            continue
        if isinstance(result, list):
            tables.extend(result)
        else:
            tables.append(result)

    return {"tables": tables}


def _parse_csv(content: bytes, filename: str = "") -> Dict[str, Any]:
    """Parse CSV — auto-detects info schema, schema dump, or data file."""
    text   = content.decode("utf-8-sig", errors="replace")
    reader = list(csv.reader(io.StringIO(text)))
    if not reader:
        return {"tables": []}

    rows_raw = [tuple(r) for r in reader]

    tbl_label = Path(filename).stem if filename else "source"
    tbl_label = re.sub(r'^[0-9a-f]{6,}-', '', tbl_label) or tbl_label

    result = _parse_sheet(tbl_label, rows_raw)
    if result is None:
        return {"tables": []}
    if isinstance(result, list):
        return {"tables": result}
    return {"tables": [result]}


def parse_schema_file(content: bytes, filename: str) -> Dict[str, Any]:
    """Return {tables: [{name, columns:[{name,type,sample,nullable}]}, ...]}"""
    from app.parsers.ddl import parse_ddl
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext in ("xlsx", "xls"):
        return _parse_excel(content)
    elif ext == "csv":
        return _parse_csv(content, filename)
    elif ext in ("sql", "ddl", "txt"):
        return parse_ddl(content.decode("utf-8-sig", errors="replace"))
    else:
        raise ValueError(f"Unsupported file type: .{ext}")
