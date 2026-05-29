"""
app/intelligence/report.py — the Mapping Report.

A single ``ReportSpec`` assembled from data the platform already captures
(mappings, readiness, lineage, governance/audit). Every renderer (JSON, HTML,
XLSX, …) is a thin transform of that one spec, so the deliverables can never
disagree. This is the "migrate with confidence before ETL" sign-off artifact
that replaces raw SQL DDL export.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import zipfile
from datetime import datetime, timezone
from typing import Dict, List, Optional

from app.intelligence import lineage as _lineage
from app.intelligence import migration_readiness as _mr


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _is_active(m: dict) -> bool:
    return (m.get("status") or "").lower() not in ("no_mapping", "skipped", "ignored", "rejected")


def build_report_spec(session: dict, src_platform: str = "generic",
                      tgt_platform: str = "generic",
                      audit_events: Optional[List[dict]] = None) -> Dict:
    """Assemble the canonical ReportSpec for a session."""
    sid = session.get("id", "")
    mappings = session.get("mappings", []) or []
    active = [m for m in mappings if _is_active(m)]
    approved = [m for m in active if (m.get("status") or "").lower() in ("approved", "accepted", "confirmed")]

    readiness = _mr.assess_session(mappings, src_platform, tgt_platform)
    lin = _lineage.build_lineage(mappings)

    versions = session.get("mapping_versions", []) or []
    sess_audit = [e for e in (audit_events or []) if e.get("session_id") == sid]
    approvals = [e for e in sess_audit
                 if any(k in (e.get("event", "") or "") for k in ("approve", "gate", "reconcile"))]

    return {
        "meta": {
            "session_id": sid,
            "name": session.get("filename") or session.get("name") or sid[:8],
            "tenant": session.get("tenant", ""),
            "created_at": session.get("created_at", ""),
            "generated_at": _now(),
            "source_platform": readiness["source_platform"],
            "target_platform": readiness["target_platform"],
        },
        "summary": {
            "total_mappings": len(mappings),
            "active_mappings": len(active),
            "approved_mappings": len(approved),
            "overall_readiness": readiness["overall_readiness"],
            "overall_level": readiness["overall_level"],
            "blockers": len(readiness["blockers"]),
            "source_tables": lin["stats"]["source_tables"],
            "target_tables": lin["stats"]["target_tables"],
            "source_columns": lin["stats"]["source_columns"],
            "target_columns": lin["stats"]["target_columns"],
        },
        "readiness": readiness,
        "lineage": lin,
        "mappings": [
            {
                "src_table": m.get("src_table", ""), "src_field": m.get("src_field", ""),
                "src_type": m.get("src_type", ""), "tgt_table": m.get("tgt_table", ""),
                "tgt_column": m.get("tgt_column", ""), "tgt_type": m.get("tgt_type", ""),
                "mapping_type": m.get("mapping_type", ""), "confidence": m.get("confidence"),
                "status": m.get("status", ""), "business_logic": m.get("business_logic", "") or "",
            }
            for m in mappings
        ],
        "risks": readiness["risks"],
        "blockers": readiness["blockers"],
        "governance": {
            "versions": len(versions),
            "approval_events": approvals[-50:],
            "audit_events": sess_audit[-100:],
            "audit_count": len(sess_audit),
        },
    }


# ── Renderers ──────────────────────────────────────────────────────────────────
_LEVEL_COLOR = {"ready": "#10b981", "review": "#0ea5e9", "risk": "#f59e0b", "blocker": "#ef4444"}


def _esc(s) -> str:
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;"))


def render_html(spec: Dict) -> str:
    m, s = spec["meta"], spec["summary"]
    lvl_color = _LEVEL_COLOR.get(s["overall_level"], "#64748b")

    def row(mp):
        c = _LEVEL_COLOR.get(_level_of(mp), "#64748b")
        conf = f"{round(mp['confidence']*100)}%" if isinstance(mp.get("confidence"), (int, float)) else "—"
        return (f"<tr><td class=m>{_esc(mp['src_table'])}.{_esc(mp['src_field'])}</td>"
                f"<td class=mut>{_esc(mp['src_type'])}</td>"
                f"<td class=m>{_esc(mp['tgt_table'])}.{_esc(mp['tgt_column'])}</td>"
                f"<td>{_esc(mp['mapping_type'])}</td><td>{conf}</td>"
                f"<td><span class=st style='color:{c}'>{_esc(mp['status'] or '—')}</span></td>"
                f"<td class=mut>{_esc(mp['business_logic'] or '')}</td></tr>")

    risk_items = "".join(f"<li><b>{_esc(r['column'])}</b> — {_esc(r['risk'])}</li>" for r in spec["risks"][:60]) \
        or "<li class=mut>No risks flagged.</li>"
    rows = "".join(row(mp) for mp in spec["mappings"]) or "<tr><td colspan=7 class=mut>No mappings.</td></tr>"
    counts = spec["readiness"]["counts"]

    return f"""<!doctype html><html><head><meta charset=utf-8>
<title>Mapping Report — {_esc(m['name'])}</title>
<style>
  body{{font-family:Inter,system-ui,sans-serif;background:#f7f8fa;color:#15181c;margin:0;padding:32px;line-height:1.5}}
  .wrap{{max-width:1000px;margin:0 auto}}
  h1{{font-size:24px;margin:0 0 2px}} .sub{{color:#6b7280;font-size:13px;margin-bottom:24px}}
  .cards{{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:24px}}
  .card{{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:16px 20px;min-width:130px}}
  .card .v{{font-size:26px;font-weight:700}} .card .l{{color:#6b7280;font-size:11px;text-transform:uppercase;letter-spacing:.05em}}
  .pill{{display:inline-block;padding:4px 12px;border-radius:20px;font-weight:700;font-size:13px;color:#fff;background:{lvl_color}}}
  h2{{font-size:15px;margin:28px 0 10px;border-bottom:1px solid #e5e7eb;padding-bottom:6px}}
  table{{width:100%;border-collapse:collapse;font-size:12.5px;background:#fff;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden}}
  th{{text-align:left;background:#f1f3f5;padding:8px 10px;font-size:11px;text-transform:uppercase;color:#6b7280}}
  td{{padding:8px 10px;border-top:1px solid #f0f1f3}}
  .m{{font-family:'JetBrains Mono',monospace;font-size:11.5px}} .mut{{color:#6b7280}} .st{{font-weight:600}}
  ul{{font-size:13px}} .foot{{color:#9aa3af;font-size:11px;margin-top:28px}}
</style></head><body><div class=wrap>
  <h1>Mapping Report — {_esc(m['name'])}</h1>
  <div class=sub>{_esc(m['source_platform'])} → {_esc(m['target_platform'])} · tenant {_esc(m['tenant'])} · generated {_esc(m['generated_at'][:19])}</div>
  <div class=cards>
    <div class=card><div class=l>Overall readiness</div><div class=v>{s['overall_readiness']}</div><span class=pill>{_esc(s['overall_level'])}</span></div>
    <div class=card><div class=l>Mappings</div><div class=v>{s['active_mappings']}</div><div class=mut>{s['approved_mappings']} approved</div></div>
    <div class=card><div class=l>Blockers</div><div class=v>{s['blockers']}</div></div>
    <div class=card><div class=l>Tables</div><div class=v>{s['source_tables']}→{s['target_tables']}</div></div>
    <div class=card><div class=l>Columns</div><div class=v>{s['source_columns']}→{s['target_columns']}</div></div>
  </div>
  <h2>Readiness breakdown</h2>
  <div class=sub>Ready {counts.get('ready',0)} · Review {counts.get('review',0)} · Risk {counts.get('risk',0)} · Blocker {counts.get('blocker',0)}</div>
  <h2>Risks &amp; blockers</h2><ul>{risk_items}</ul>
  <h2>Mapping specification</h2>
  <table><thead><tr><th>Source</th><th>Type</th><th>Target</th><th>Match</th><th>Conf.</th><th>Status</th><th>Transform</th></tr></thead>
  <tbody>{rows}</tbody></table>
  <h2>Governance</h2>
  <div class=sub>Versions captured: {spec['governance']['versions']} · Audit events: {spec['governance']['audit_count']} · Approval events: {len(spec['governance']['approval_events'])}</div>
  <div class=foot>Generated by xREF DataMapper. This report is a migration decision artifact — feed the mapping specification (CSV/XLSX) into your ETL/orchestration tooling.</div>
</div></body></html>"""


def _level_of(mp: dict) -> str:
    """Best-effort level for a mapping row from its confidence (HTML coloring)."""
    c = mp.get("confidence")
    if isinstance(c, (int, float)):
        if c >= 0.9:
            return "ready"
        if c >= 0.75:
            return "review"
        if c >= 0.6:
            return "risk"
        return "blocker"
    return "review"


def render_xlsx(spec: Dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    hdr_font = Font(bold=True, color="FFFFFF")

    # Summary
    ws = wb.active
    ws.title = "Summary"
    m, s = spec["meta"], spec["summary"]
    ws.append(["Mapping Report", m["name"]])
    ws["A1"].font = Font(bold=True, size=14)
    for k, v in [
        ("Source platform", m["source_platform"]), ("Target platform", m["target_platform"]),
        ("Tenant", m["tenant"]), ("Generated", m["generated_at"][:19]),
        ("Overall readiness", s["overall_readiness"]), ("Readiness level", s["overall_level"]),
        ("Active mappings", s["active_mappings"]), ("Approved", s["approved_mappings"]),
        ("Blockers", s["blockers"]),
        ("Tables (src→tgt)", f"{s['source_tables']}→{s['target_tables']}"),
        ("Columns (src→tgt)", f"{s['source_columns']}→{s['target_columns']}"),
    ]:
        ws.append([k, v])
        ws.cell(ws.max_row, 1).font = bold

    # Mappings
    wm = wb.create_sheet("Mappings")
    cols = ["src_table", "src_field", "src_type", "tgt_table", "tgt_column", "tgt_type",
            "mapping_type", "confidence", "status", "business_logic"]
    wm.append([c.replace("_", " ").title() for c in cols])
    for c_i in range(1, len(cols) + 1):
        wm.cell(1, c_i).fill = hdr_fill
        wm.cell(1, c_i).font = hdr_font
    for mp in spec["mappings"]:
        wm.append([mp.get(c, "") for c in cols])

    # Risks
    wr = wb.create_sheet("Risks")
    wr.append(["Column", "Risk"])
    for c_i in (1, 2):
        wr.cell(1, c_i).fill = hdr_fill
        wr.cell(1, c_i).font = hdr_font
    for r in spec["risks"]:
        wr.append([r.get("column", ""), r.get("risk", "")])

    # Governance
    wg = wb.create_sheet("Governance")
    wg.append(["Timestamp", "Event", "User", "Detail"])
    for c_i in range(1, 5):
        wg.cell(1, c_i).fill = hdr_fill
        wg.cell(1, c_i).font = hdr_font
    for e in spec["governance"]["audit_events"]:
        wg.append([e.get("ts", ""), e.get("event", ""), e.get("email", ""), str(e.get("meta", ""))])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── STM mapping CSV ─────────────────────────────────────────────────────────────
_CSV_COLS = ["src_table", "src_field", "src_type", "tgt_table", "tgt_column",
             "tgt_type", "mapping_type", "business_logic", "confidence", "status"]


def render_mappings_csv(spec: Dict) -> str:
    """Source-to-target-mapping (STM) CSV — the ETL-feed format."""
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=_CSV_COLS, extrasaction="ignore")
    w.writeheader()
    for mp in spec["mappings"]:
        w.writerow({k: mp.get(k, "") for k in _CSV_COLS})
    return buf.getvalue()


# ── Verification hash (tamper-evidence for the certificate) ─────────────────────
def verification_hash(spec: Dict) -> str:
    """Deterministic digest over the material facts, so a certificate can be
    verified against the report it was issued from."""
    m, s = spec["meta"], spec["summary"]
    material = {
        "session_id": m["session_id"],
        "source": m["source_platform"], "target": m["target_platform"],
        "overall_readiness": s["overall_readiness"], "level": s["overall_level"],
        "active": s["active_mappings"], "approved": s["approved_mappings"],
        "blockers": s["blockers"],
        "columns": sorted(f"{mp['src_table']}.{mp['src_field']}->{mp['tgt_table']}.{mp['tgt_column']}"
                          for mp in spec["mappings"]),
    }
    return hashlib.sha256(json.dumps(material, sort_keys=True).encode()).hexdigest()


# ── Migration Readiness Certificate (PDF) ───────────────────────────────────────
def render_certificate_pdf(spec: Dict) -> bytes:
    """One-page, branded readiness certificate — the migration sign-off artifact."""
    from fpdf import FPDF  # pure-python, no system deps

    def _l1(x):  # core fonts are Latin-1 only; sanitize user/dynamic text
        return (str(x).replace("→", "->").replace("—", "-").replace("–", "-")
                .replace("’", "'").replace("•", "-")
                .encode("latin-1", "replace").decode("latin-1"))

    m, s = spec["meta"], spec["summary"]
    lvl = s["overall_level"]
    rgb = {"ready": (16, 185, 129), "review": (14, 165, 233),
           "risk": (245, 158, 11), "blocker": (239, 68, 68)}.get(lvl, (100, 116, 139))
    approved_pct = round(100 * s["approved_mappings"] / s["active_mappings"]) if s["active_mappings"] else 0
    vhash = verification_hash(spec)

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(False)
    pdf.add_page()

    # Accent header bar
    pdf.set_fill_color(*rgb)
    pdf.rect(0, 0, 210, 6, style="F")

    pdf.set_xy(15, 22)
    pdf.set_text_color(20, 24, 28)
    pdf.set_font("Helvetica", "B", 22)
    pdf.cell(0, 10, "Migration Readiness Certificate")
    pdf.set_xy(15, 34)
    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(110, 116, 128)
    pdf.cell(0, 8, _l1(f"{m['name']}   |   {m['source_platform']} -> {m['target_platform']}"))

    # Score block
    pdf.set_fill_color(248, 249, 251)
    pdf.rect(15, 50, 180, 34, style="F")
    pdf.set_xy(22, 55)
    pdf.set_text_color(*rgb)
    pdf.set_font("Helvetica", "B", 40)
    pdf.cell(40, 24, str(s["overall_readiness"]))
    pdf.set_xy(62, 58)
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 8, f"{lvl.upper()}")
    pdf.set_xy(62, 68)
    pdf.set_text_color(110, 116, 128)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, "overall readiness score (0-100)")

    # Gates
    gates = [
        ("Blockers", str(s["blockers"]), s["blockers"] == 0),
        ("Approved", f"{approved_pct}%", approved_pct == 100),
        ("Mappings", str(s["active_mappings"]), s["active_mappings"] > 0),
        ("Versions", str(spec["governance"]["versions"]), True),
    ]
    x = 15
    for label, val, ok in gates:
        pdf.set_fill_color(245, 247, 249)
        pdf.rect(x, 92, 42, 24, style="F")
        pdf.set_xy(x, 96)
        pdf.set_text_color(20, 24, 28)
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(42, 8, val, align="C")
        pdf.set_xy(x, 106)
        pdf.set_text_color(120, 130, 140)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(42, 6, f"{label} {'PASS' if ok else 'CHECK'}", align="C")
        x += 46

    # Sign-off
    pdf.set_xy(15, 128)
    pdf.set_text_color(20, 24, 28)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Sign-off")
    approvals = spec["governance"]["approval_events"]
    last_by = approvals[-1].get("email", "") if approvals else ""
    last_ts = (approvals[-1].get("ts", "") if approvals else "")[:19]
    y = 138
    for role in ("Analyst", "Architect", "Lead"):
        pdf.set_xy(15, y)
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(110, 116, 128)
        pdf.cell(30, 8, role + ":")
        pdf.set_draw_color(200, 205, 212)
        pdf.line(45, y + 7, 120, y + 7)
        y += 12
    if last_by:
        pdf.set_xy(125, 138)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(110, 116, 128)
        pdf.multi_cell(70, 5, _l1(f"Last approval event:\n{last_by}\n{last_ts}"))

    # Verification + footer
    pdf.set_xy(15, 182)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(140, 146, 156)
    pdf.multi_cell(180, 4,
                   f"Verification hash (SHA-256): {vhash}\n"
                   f"Issued: {m['generated_at'][:19]}  |  Tenant: {m['tenant']}  |  Session: {m['session_id']}")
    pdf.set_xy(15, 280)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 4, "Generated by xREF DataMapper - migration decision artifact. Verify the hash against the source report.")

    out = pdf.output()
    return bytes(out)


# ── Change-management bundle (ZIP) ──────────────────────────────────────────────
def build_bundle_zip(spec: Dict, include_pdf: bool = True) -> bytes:
    """A single ZIP an architect can hand to their Change Advisory Board:
    report + mapping CSV/XLSX + readiness certificate + audit log + lineage."""
    sid = spec["meta"]["session_id"][:8]
    vhash = verification_hash(spec)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(f"mapping_report_{sid}.html", render_html(spec))
        z.writestr(f"mapping_report_{sid}.xlsx", render_xlsx(spec))
        z.writestr(f"mappings_{sid}.csv", render_mappings_csv(spec))
        z.writestr(f"audit_log_{sid}.json",
                   json.dumps(spec["governance"]["audit_events"], indent=2))
        z.writestr(f"lineage_{sid}.json", json.dumps(spec["lineage"], indent=2))
        if include_pdf:
            try:
                z.writestr(f"readiness_certificate_{sid}.pdf", render_certificate_pdf(spec))
            except Exception:
                pass  # fpdf unavailable — bundle still useful without the cert
        manifest = (
            "xREF DataMapper — Change-Management Bundle\n"
            f"Session:  {spec['meta']['session_id']}\n"
            f"System:   {spec['meta']['name']} ({spec['meta']['source_platform']} -> {spec['meta']['target_platform']})\n"
            f"Readiness:{spec['summary']['overall_readiness']} ({spec['summary']['overall_level']})  "
            f"Blockers: {spec['summary']['blockers']}\n"
            f"Issued:   {spec['meta']['generated_at']}\n"
            f"Verify:   SHA-256 {vhash}\n\n"
            "Contents:\n"
            "  mapping_report_*.html   — full report (open in a browser)\n"
            "  mapping_report_*.xlsx   — report workbook\n"
            "  mappings_*.csv          — STM spec (feed your ETL/orchestration)\n"
            "  readiness_certificate_*.pdf — one-page sign-off certificate\n"
            "  audit_log_*.json        — governance/audit events\n"
            "  lineage_*.json          — source->target lineage graph\n"
        )
        z.writestr("README.txt", manifest)
    return buf.getvalue()
