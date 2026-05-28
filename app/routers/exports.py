"""
app/routers/exports.py — /api/sessions/{sid}/export/* + /api/sessions/{sid}/sql
"""
from __future__ import annotations

import csv
import io
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from typing import Dict, List

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse

from app.core.audit import _now, _write_audit_event
from app.core.session_store import _session_or_404
from app.intelligence.sql_format import (
    DIALECT_RENDERERS,
    render_sql_for_dialect,
)
from app.routers._helpers import _get_client_ip

router = APIRouter()


@router.get("/api/sessions/{sid}/export/csv")
async def export_csv(sid: str, request: Request):
    s = _session_or_404(sid)
    mappings = s.get("mappings", [])
    _write_audit_event("export.csv", tenant=s.get("tenant"), session_id=sid,
                       ip=_get_client_ip(request), metadata={"rows": len(mappings)})
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=[
        "src_table", "src_field", "src_type",
        "tgt_table", "tgt_column", "tgt_type",
        "mapping_type", "business_logic",
        "confidence", "tier", "status", "rationale",
    ])
    w.writeheader()
    for m in mappings:
        w.writerow({k: m.get(k, "") for k in w.fieldnames})
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="stm_{sid[:8]}.csv"'},
    )


@router.get("/api/sessions/{sid}/export/xlsx")
async def export_xlsx(sid: str):
    s = _session_or_404(sid)
    mappings = s.get("mappings", [])

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "STM Mappings"

    HEADERS = [
        "Src Table", "Src Field", "Src Type", "→",
        "Tgt Table", "Tgt Column", "Tgt Type",
        "Mapping Type", "Business Logic",
        "Confidence %", "Tier", "Status", "PII Class", "Rationale",
    ]
    HDR_FILL = PatternFill("solid", fgColor="141413")
    HDR_FONT = Font(bold=True, color="FAFAFA", name="Calibri", size=10)
    TIER_FILL = {
        "high":   PatternFill("solid", fgColor="EEF4E8"),
        "medium": PatternFill("solid", fgColor="FEF6EC"),
        "low":    PatternFill("solid", fgColor="FDF0F0"),
        "none":   PatternFill("solid", fgColor="F2F0E8"),
    }

    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, m in enumerate(mappings, 2):
        values = [
            m.get("src_table", ""),  m.get("src_field", ""),  m.get("src_type", ""),
            "→" if m.get("tgt_column") else "✕",
            m.get("tgt_table", ""),  m.get("tgt_column", ""), m.get("tgt_type", ""),
            m.get("mapping_type", ""),
            m.get("business_logic", ""),
            round((m.get("confidence") or 0) * 100),
            m.get("tier", ""),  m.get("status", ""),
            m.get("pii_class", "auto"),  m.get("rationale", ""),
        ]
        fill = TIER_FILL.get(m.get("tier", ""), PatternFill("solid", fgColor="FFFFFF"))
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill

    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col) + 4
        ws.column_dimensions[col[0].column_letter].width = min(w, 52)
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")
    stats = s.get("stats", {})
    ws2["A1"] = "xREF Agent — Session Summary"
    ws2["A1"].font = Font(bold=True, size=14, name="Calibri")
    rows2 = [
        ("Session ID",     sid),
        ("Generated At",   _now()),
        ("Source File",    s.get("filename", "")),
        ("",               ""),
        ("Total Fields",   stats.get("total", 0)),
        ("Auto-Mapped",    stats.get("mapped", 0)),
        ("Needs Review",   stats.get("review", 0)),
        ("Unmapped",       stats.get("unmapped", 0)),
        ("Avg Confidence", f"{round((stats.get('avg_confidence') or 0) * 100)}%"),
    ]
    for i, (k, v) in enumerate(rows2, 3):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True, name="Calibri")
        ws2.cell(row=i, column=2, value=v)
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 24

    ws3 = wb.create_sheet("Metadata")
    ws3["A1"] = "xREF Agent — Mapping Provenance"
    ws3["A1"].font = Font(bold=True, size=13, name="Calibri", color="1A56DB")

    src_tables_list = sorted({m.get("src_table", "") for m in mappings if m.get("src_table")})
    tgt_tables_list = sorted({m.get("tgt_table", "") for m in mappings if m.get("tgt_table")})

    ws3.cell(row=3, column=1, value="Source Tables").font = Font(bold=True, name="Calibri")
    for i, t in enumerate(src_tables_list, 4):
        ws3.cell(row=i, column=1, value=t)

    tgt_start = 3 + len(src_tables_list) + 2
    ws3.cell(row=tgt_start, column=1, value="Target Tables").font = Font(bold=True, name="Calibri")
    for i, t in enumerate(tgt_tables_list, tgt_start + 1):
        ws3.cell(row=i, column=1, value=t)

    ws3.cell(row=3, column=3, value="Mapping Type").font = Font(bold=True, name="Calibri")
    ws3.cell(row=3, column=4, value="Count").font = Font(bold=True, name="Calibri")
    type_counts = Counter(m.get("mapping_type", "Unknown") for m in mappings)
    for i, (mtype, cnt) in enumerate(sorted(type_counts.items()), 4):
        ws3.cell(row=i, column=3, value=mtype)
        ws3.cell(row=i, column=4, value=cnt)

    ws3.cell(row=3, column=6, value="Confidence Tier").font = Font(bold=True, name="Calibri")
    ws3.cell(row=3, column=7, value="Count").font = Font(bold=True, name="Calibri")
    tier_counts = Counter(m.get("tier", "none") for m in mappings)
    for i, (tier, cnt) in enumerate(sorted(tier_counts.items()), 4):
        ws3.cell(row=i, column=6, value=tier.capitalize())
        ws3.cell(row=i, column=7, value=cnt)

    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 26

    ws4 = wb.create_sheet("Table Mapping Summary")
    ws4["A1"] = "xREF Agent — Table-Level Mapping Summary"
    ws4["A1"].font = Font(bold=True, size=13, name="Calibri", color="1A56DB")

    TBL_HEADERS = [
        "Source Table", "Target Table",
        "Total Cols", "Mapped", "Review", "Unmapped",
        "Coverage %", "Avg Confidence %",
        "Mapping Types", "Relation Pattern",
    ]
    thin = Side(style="thin", color="CBD5E1")
    border = Border(bottom=thin)

    for ci, h in enumerate(TBL_HEADERS, 1):
        cell = ws4.cell(row=3, column=ci, value=h)
        cell.font = Font(bold=True, name="Calibri", size=9, color="334155")
        cell.fill = PatternFill("solid", fgColor="F1F5F9")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    pair_stats: Dict = defaultdict(lambda: {
        "total": 0, "mapped": 0, "review": 0, "unmapped": 0,
        "conf_sum": 0.0, "conf_count": 0,
        "types": set(), "relations": set(),
    })
    for m in mappings:
        src_t = m.get("src_table", "(unknown)")
        tgt_t = m.get("tgt_table", "") or "(unmapped)"
        key   = (src_t, tgt_t)
        ps    = pair_stats[key]
        ps["total"] += 1
        st = m.get("status", "unmapped")
        if st == "mapped":
            ps["mapped"] += 1
        elif st == "review":
            ps["review"] += 1
        else:
            ps["unmapped"] += 1
        if m.get("confidence"):
            ps["conf_sum"]   += float(m["confidence"])
            ps["conf_count"] += 1
        if m.get("mapping_type"):
            ps["types"].add(m["mapping_type"])
        if m.get("mapping_relation"):
            ps["relations"].add(m["mapping_relation"])

    ROW_FILLS = [PatternFill("solid", fgColor="FFFFFF"), PatternFill("solid", fgColor="F8FAFC")]
    for ri, ((src_t, tgt_t), ps) in enumerate(sorted(pair_stats.items()), 4):
        total    = ps["total"]
        cov      = round(ps["mapped"] / max(total, 1) * 100, 1)
        avg_conf = round(ps["conf_sum"] / max(ps["conf_count"], 1) * 100, 1)
        row_fill = ROW_FILLS[ri % 2]
        values   = [
            src_t, tgt_t,
            total, ps["mapped"], ps["review"], ps["unmapped"],
            f"{cov}%", f"{avg_conf}%",
            ", ".join(sorted(ps["types"])),
            ", ".join(sorted(ps["relations"])),
        ]
        for ci, val in enumerate(values, 1):
            cell = ws4.cell(row=ri, column=ci, value=val)
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=9,
                             color="059669" if ci == 7 and cov >= 80
                             else ("DC2626" if ci == 7 and cov < 60 else "0F172A"))
            cell.alignment = Alignment(vertical="center")

    ws4_col_widths = [28, 28, 10, 10, 10, 10, 12, 16, 30, 20]
    for ci, w in enumerate(ws4_col_widths, 1):
        ws4.column_dimensions[ws4.cell(row=3, column=ci).column_letter].width = w
    ws4.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="stm_{sid[:8]}.xlsx"'},
    )


def _group_mappings_by_target(mappings: List[dict]) -> Dict[str, List[dict]]:
    """Group mapping rows by their target table, skipping unmapped rows."""
    grouped: Dict[str, List[dict]] = defaultdict(list)
    for m in mappings:
        tgt_t = (m.get("tgt_table") or "").strip()
        if not tgt_t:
            continue
        grouped[tgt_t].append(m)
    return grouped


def _safe_filename(name: str) -> str:
    """Normalise a target-table name into a safe filename stem."""
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", name).strip("._")
    return cleaned or "table"


@router.get("/api/sessions/{sid}/export/sql")
async def export_sql(sid: str, dialect: str = "bigquery"):
    """Render the session's SQL for a target dialect.

    Supported dialects: bigquery (MERGE), snowflake (MERGE),
    spark (INSERT OVERWRITE), ansi (INSERT INTO ... SELECT).

    For backwards compatibility, when ``dialect=bigquery`` and the session
    already has a pre-generated CREATE-OR-REPLACE block in ``generated_sql``
    we return that verbatim. Otherwise we render per-target blocks via
    ``render_sql_for_dialect``.
    """
    s = _session_or_404(sid)
    dialect_norm = (dialect or "bigquery").lower()
    if dialect_norm not in DIALECT_RENDERERS:
        dialect_norm = "ansi"

    # Preserve legacy behaviour: a pre-generated BigQuery block wins.
    if dialect_norm == "bigquery" and s.get("generated_sql"):
        sql = s["generated_sql"]
        return StreamingResponse(
            io.BytesIO(sql.encode()),
            media_type="text/plain",
            headers={"Content-Disposition": f'attachment; filename="stm_{sid[:8]}.sql"'},
        )

    mappings = s.get("mappings", [])
    grouped = _group_mappings_by_target(mappings)
    if not grouped:
        raise HTTPException(
            422,
            "No mapped target columns available. Complete the pipeline first.",
        )

    header = (
        f"-- xREF Agent — Dialect: {dialect_norm}\n"
        f"-- Session: {sid}\n"
        f"-- Generated: {_now()}\n"
        f"-- Target tables: {len(grouped)}\n\n"
    )
    blocks: List[str] = [header]
    for tbl in sorted(grouped.keys()):
        rows = grouped[tbl]
        block_sql = render_sql_for_dialect(dialect_norm, tbl, rows)
        blocks.append(f"-- ── Target table: {tbl} ({len(rows)} cols) ──\n")
        blocks.append(block_sql)
        blocks.append("\n")
    sql = "".join(blocks)

    return StreamingResponse(
        io.BytesIO(sql.encode()),
        media_type="text/plain",
        headers={
            "Content-Disposition": (
                f'attachment; filename="stm_{sid[:8]}_{dialect_norm}.sql"'
            )
        },
    )


@router.get("/api/sessions/{sid}/export/sql/zip")
async def export_sql_zip(sid: str, dialect: str = "bigquery"):
    """Return one .sql file per target table, zipped together."""
    s = _session_or_404(sid)
    dialect_norm = (dialect or "bigquery").lower()
    if dialect_norm not in DIALECT_RENDERERS:
        dialect_norm = "ansi"

    mappings = s.get("mappings", [])
    grouped = _group_mappings_by_target(mappings)
    if not grouped:
        raise HTTPException(
            422,
            "No mapped target columns available. Complete the pipeline first.",
        )

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        readme_lines = [
            "xREF Agent — per-table SQL bundle",
            f"Session: {sid}",
            f"Dialect: {dialect_norm}",
            f"Generated: {_now()}",
            f"Target tables: {len(grouped)}",
            "",
            "Files:",
        ]
        for tbl in sorted(grouped.keys()):
            rows = grouped[tbl]
            sql_block = (
                f"-- xREF Agent — Dialect: {dialect_norm}\n"
                f"-- Target table: {tbl} ({len(rows)} cols)\n"
                f"-- Session: {sid}\n"
                f"-- Generated: {_now()}\n\n"
                f"{render_sql_for_dialect(dialect_norm, tbl, rows)}"
            )
            fname = f"{_safe_filename(tbl)}.sql"
            zf.writestr(fname, sql_block)
            readme_lines.append(f"  - {fname} ({len(rows)} cols)")
        zf.writestr("README.txt", "\n".join(readme_lines) + "\n")

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={
            "Content-Disposition": (
                f'attachment; filename="stm_{sid[:8]}_{dialect_norm}_sql.zip"'
            )
        },
    )


@router.get("/api/sessions/{sid}/export/table-mappings")
async def export_table_mappings_csv(sid: str):
    """Download a CSV of per table-pair mapping statistics."""
    s = _session_or_404(sid)
    mappings = s.get("mappings", [])
    if not mappings:
        raise HTTPException(422, "No mappings found — run the pipeline first")

    stats: Dict = {}
    for m in mappings:
        src_t = m.get("src_table", "")
        tgt_t = m.get("tgt_table", "") or "UNMAPPED"
        if m.get("status") == "unmapped":
            tgt_t = "UNMAPPED"
        key = (src_t, tgt_t)
        if key not in stats:
            stats[key] = {
                "src_table": src_t, "tgt_table": tgt_t,
                "total": 0, "mapped": 0, "review": 0, "unmapped": 0,
                "conf_sum": 0.0, "conf_count": 0,
                "relations": set(), "types": set(),
            }
        ps = stats[key]
        ps["total"] += 1
        status = m.get("status", "unmapped")
        ps[status] = ps.get(status, 0) + 1
        if m.get("confidence", 0) > 0:
            ps["conf_sum"]   += m["confidence"]
            ps["conf_count"] += 1
        if m.get("mapping_relation"):
            ps["relations"].add(m["mapping_relation"])
        if m.get("mapping_type") and m.get("mapping_type") != "Unused":
            ps["types"].add(m["mapping_type"])

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow([
        "src_table", "tgt_table", "relation", "total_cols",
        "mapped", "review", "unmapped", "coverage_pct",
        "avg_confidence", "mapping_types", "notes",
    ])

    sorted_pairs = sorted(
        stats.values(),
        key=lambda r: (-round((r["mapped"] + r["review"]) / max(r["total"], 1) * 100, 1), r["src_table"]),
    )

    for r in sorted_pairs:
        total = r["total"]
        mapped = r.get("mapped", 0)
        review = r.get("review", 0)
        unmapped = r.get("unmapped", 0)
        covered = mapped + review
        coverage_pct = round(covered / max(total, 1) * 100, 1)
        avg_conf = round(r["conf_sum"] / max(r["conf_count"], 1), 3) if r["conf_count"] else 0.0
        relations = "|".join(sorted(r["relations"])) if r["relations"] else "1:1"
        types_str = "|".join(sorted(r["types"])) if r["types"] else "Direct"

        flags = []
        if coverage_pct == 100.0:
            flags.append("COMPLETE")
        if review > 0:
            flags.append("NEEDS REVIEW")
        if unmapped > 0 and (unmapped / max(total, 1)) > 0.20:
            flags.append(f"HIGH UNMAPPED ({unmapped}/{total})")
        notes = "; ".join(flags) if flags else ""

        writer.writerow([
            r["src_table"], r["tgt_table"], relations, total,
            mapped, review, unmapped,
            f"{coverage_pct}%", avg_conf, types_str, notes,
        ])

    csv_bytes = out.getvalue().encode("utf-8")
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="table_mappings_{sid[:8]}_{ts}.csv"',
            "Content-Length": str(len(csv_bytes)),
        },
    )


@router.get("/api/sessions/{sid}/sql")
async def get_sql(sid: str):
    s = _session_or_404(sid)
    return {"sql": s.get("generated_sql", ""), "ready": bool(s.get("generated_sql"))}
